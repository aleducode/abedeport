#!/usr/bin/env python3
"""
Production SQL generator for tournament data
Usage: python3 generate-sql-prod.py

This script reads Excel files from tablas/ directory and generates SQL INSERT statements
for production environment with stronger security considerations.
"""

import os
import sys
from openpyxl import load_workbook
import re

def sanitize_team_name(name):
    """Sanitize team name for SQL security"""
    if not name:
        return "Unknown Team"
    
    # Remove or replace potentially problematic characters
    name = str(name).strip()
    
    # Replace single quotes with escaped version
    name = name.replace("'", "\\'")
    
    # Remove any SQL injection attempts (basic protection)
    dangerous_patterns = ['--', ';', 'DROP', 'DELETE', 'INSERT', 'UPDATE', 'SELECT']
    for pattern in dangerous_patterns:
        name = re.sub(pattern, '', name, flags=re.IGNORECASE)
    
    # Limit length
    name = name[:100]
    
    return name if name else "Unknown Team"

def read_excel_tournament(file_path, sport):
    """Read tournament data from Excel file with enhanced validation"""
    print(f"📖 Reading {file_path}...")
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        teams = []
        
        # Find header row (look for common column headers)
        header_row = None
        for row_num in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_values = [str(cell.value).lower() if cell.value else '' for cell in sheet[row_num]]
            row_text = ' '.join(row_values)
            
            # Look for typical tournament table headers
            if any(keyword in row_text for keyword in ['pos', 'equipo', 'team', 'pj', 'pg', 'pp', 'pe', 'puntos']):
                header_row = row_num
                break
        
        if not header_row:
            print(f"⚠️  Could not find header row in {file_path}")
            return []
        
        print(f"Found header row at line {header_row}")
        
        # Read team data starting from the row after header
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Skip empty rows
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                continue
            
            # Extract basic data - adjust column indices based on common Excel formats
            values = [cell.value for cell in row]
            
            # Try to identify position and team name
            pos = None
            team_name = None
            pj = pg = pp = pe = pf = pc = pts = 0
            
            # Look for position (usually first numeric column)
            for i, val in enumerate(values):
                if isinstance(val, (int, float)) and val > 0 and val <= 50:  # Reasonable position range
                    pos = int(val)
                    break
            
            # Look for team name (usually first text column after position)
            for i, val in enumerate(values):
                if isinstance(val, str) and val.strip() and not val.strip().isdigit():
                    team_name = sanitize_team_name(val.strip())
                    break
            
            # Extract numeric stats (PJ, PG, PP, PE, PF, PC, Pts)
            numeric_values = []
            for val in values:
                if isinstance(val, (int, float)) and val != pos:  # Exclude position
                    # Validate numeric ranges
                    num_val = int(val)
                    if -1000 <= num_val <= 1000:  # Reasonable range for sports stats
                        numeric_values.append(num_val)
            
            # Map numeric values to stats (common order: PJ, PG, PP, PE, PF, PC, Pts)
            if len(numeric_values) >= 3:
                pj = max(0, numeric_values[0]) if len(numeric_values) > 0 else 0
                pg = max(0, numeric_values[1]) if len(numeric_values) > 1 else 0
                pp = max(0, numeric_values[2]) if len(numeric_values) > 2 else 0
                pe = max(0, numeric_values[3]) if len(numeric_values) > 3 else 0
                pf = numeric_values[4] if len(numeric_values) > 4 else 0
                pc = numeric_values[5] if len(numeric_values) > 5 else 0
                pts = numeric_values[6] if len(numeric_values) > 6 else max(0, pg * 3 + pe)  # Calculate if not provided
            
            # Only add if we have essential data and valid ranges
            if pos and team_name and 1 <= pos <= 50 and pj >= 0:
                teams.append({
                    'posicion': pos,
                    'nombre_equipo': team_name,
                    'partidos_jugados': pj,
                    'partidos_ganados': pg,
                    'partidos_perdidos': pp,
                    'partidos_empatados': pe,
                    'puntos_favor': pf,
                    'puntos_contra': pc,
                    'puntos_totales': pts
                })
        
        # Sort by position and validate no duplicates
        teams.sort(key=lambda x: x['posicion'])
        
        # Remove duplicate positions
        seen_positions = set()
        unique_teams = []
        for team in teams:
            if team['posicion'] not in seen_positions:
                seen_positions.add(team['posicion'])
                unique_teams.append(team)
        
        print(f"✅ Extracted {len(unique_teams)} teams from {sport} tournament")
        
        return unique_teams
        
    except Exception as e:
        print(f"❌ Error reading {file_path}: {str(e)}")
        return []

def generate_production_sql():
    """Generate SQL statements for production tournament data"""
    
    print("🏆 Production Tournament SQL Generator")
    print("=====================================")
    print("⚠️  Production environment - Enhanced security validation")
    print()
    
    # Check if tablas directory exists
    tablas_dir = "tablas"
    if not os.path.exists(tablas_dir):
        print("❌ tablas/ directory not found")
        return False
    
    # Tournament file mappings
    tournament_files = {
        'futsal': 'ligabfutsal.xlsx',
        'baloncesto': 'baloncesto.xlsx', 
        'voleibol': 'voleibol.xlsx'
    }
    
    tournament_names = {
        'futsal': 'Liga de Futsal 2024',
        'baloncesto': 'Liga de Baloncesto 2024',
        'voleibol': 'Liga de Voleibol 2024'
    }
    
    # Output SQL file for production
    output_file = "tournament_import_prod.sql"
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("-- Production Tournament Data Import SQL\n")
        f.write("-- Generated automatically from Excel files with enhanced security\n")
        f.write("-- Execute this file in production database\n")
        f.write("-- Recommended: Use environment variables for credentials\n\n")
        
        # Clean existing data with transaction
        f.write("-- Clean existing tournament data (with transaction safety)\n")
        f.write("START TRANSACTION;\n\n")
        f.write("SET FOREIGN_KEY_CHECKS = 0;\n")
        f.write("DELETE FROM equipos_tournament;\n")
        f.write("DELETE FROM tournaments;\n")
        f.write("ALTER TABLE equipos_tournament AUTO_INCREMENT = 1;\n")
        f.write("ALTER TABLE tournaments AUTO_INCREMENT = 1;\n")
        f.write("SET FOREIGN_KEY_CHECKS = 1;\n\n")
        
        tournament_id = 1
        total_teams = 0
        
        # Process each tournament
        for sport, filename in tournament_files.items():
            file_path = os.path.join(tablas_dir, filename)
            
            if not os.path.exists(file_path):
                print(f"⚠️  File not found: {file_path} (skipping)")
                continue
            
            # Read teams from Excel
            teams = read_excel_tournament(file_path, sport)
            
            if not teams:
                print(f"⚠️  No teams found in {filename} (skipping)")
                continue
            
            # Write tournament
            tournament_name = tournament_names[sport]
            f.write(f"-- {tournament_name} (Tournament ID: {tournament_id})\n")
            f.write(f"INSERT INTO tournaments (id_tournament, nombre, deporte, temporada, estado, fecha_inicio) VALUES\n")
            f.write(f"({tournament_id}, '{tournament_name}', '{sport}', '2024', 'activo', NOW());\n\n")
            
            # Write teams with validation comments
            f.write(f"-- Teams for {tournament_name} ({len(teams)} teams)\n")
            f.write(f"INSERT INTO equipos_tournament (id_tournament, nombre_equipo, partidos_jugados, partidos_ganados, partidos_perdidos, partidos_empatados, puntos_favor, puntos_contra, puntos_totales, posicion) VALUES\n")
            
            team_inserts = []
            for team in teams:
                team_inserts.append(
                    f"({tournament_id}, '{team['nombre_equipo']}', {team['partidos_jugados']}, {team['partidos_ganados']}, {team['partidos_perdidos']}, {team['partidos_empatados']}, {team['puntos_favor']}, {team['puntos_contra']}, {team['puntos_totales']}, {team['posicion']})"
                )
            
            f.write(",\n".join(team_inserts) + ";\n\n")
            total_teams += len(teams)
            tournament_id += 1
        
        f.write("-- Commit transaction\n")
        f.write("COMMIT;\n\n")
        f.write("-- Production import completed\n")
        f.write(f"-- Imported {tournament_id - 1} tournaments with {total_teams} teams total\n")
        f.write("-- Verify data integrity after import\n")
        f.write("SELECT 'Tournaments imported:' as info, COUNT(*) as count FROM tournaments;\n")
        f.write("SELECT 'Teams imported:' as info, COUNT(*) as count FROM equipos_tournament;\n")
    
    print(f"✅ Production SQL file generated: {output_file}")
    print(f"📊 Summary: {tournament_id - 1} tournaments, {total_teams} teams")
    print(f"🔒 Enhanced security validations applied")
    print()
    print("📋 Production import commands:")
    print("# Using environment variables (recommended):")
    print("export DB_PASSWORD='your_production_password'")
    print("docker-compose -f production.yml up -d db")
    print("docker-compose -f production.yml exec -T db mysql -u abedeport_user -p$DB_PASSWORD ABEDEPORT < tournament_import_prod.sql")
    print()
    print("# Or using default password:")
    print("docker-compose -f production.yml up -d db")
    print("docker-compose -f production.yml exec -T db mysql -u abedeport_user -pabedeport_strong_password ABEDEPORT < tournament_import_prod.sql")
    
    return True

if __name__ == "__main__":
    generate_production_sql()